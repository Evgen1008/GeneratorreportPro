from openpyxl import load_workbook


SHEETS = {
    1: "01-10",
    2: "11-20",
    3: "21-31",
}


def read_initial_balance(filename, decade=1):
    """
    Читает начальный остаток топлива
    из верхней строки исходного Excel.

    Если найдено:
    "Остаток топлива на начало" -> берём число справа.

    Если нет:
    возвращает None.
    """

    print("!!! NEW EXCEL READER !!!")

    wb = load_workbook(
        filename,
        data_only=True
    )

    sheet = SHEETS.get(
        decade,
        "01-10"
    )

    if sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        ws = wb.worksheets[0]

    for row in ws.iter_rows():

        for cell in row:

            if cell.value:

                text = str(
                    cell.value
                ).lower()

                if (
                    "остаток" in text
                    and
                    "топлив" in text
                ):

                    print(
                        "FOUND BALANCE LABEL:",
                        cell.coordinate,
                        cell.value
                    )

                    for col in range(
                        cell.column + 1,
                        ws.max_column + 1
                    ):

                        value = ws.cell(
                            cell.row,
                            col
                        ).value

                        if value not in (
                            None,
                            ""
                        ):
                            try:
                                result = float(value)

                                print(
                                    "BALANCE:",
                                    result
                                )

                                return result

                            except Exception:
                                pass

    print("BALANCE NOT FOUND")

    return None



def read_start_balance(
        filename,
        decade,
        first_work_date=None
):
    """
    Совместимость со старым кодом.
    """

    return read_initial_balance(
        filename,
        decade
    )