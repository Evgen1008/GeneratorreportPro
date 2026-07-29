import os
import shutil
import calendar
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from app.utils.paths import REPORTS_DIR, TEMPLATES_DIR


class ExcelReport:

    def __init__(self, reports_folder=None, database=None):
        self.reports_folder = reports_folder or REPORTS_DIR
        self.db = database

    def get_decade_sheet(self, decade):
        return {1: "01-10", 2: "11-20", 3: "21-31"}.get(decade, "01-10")

    def set_cell(self, ws, cell, value):
        if not isinstance(ws[cell], MergedCell):
            ws[cell] = value

    def clear_rows(self, ws):
        for row in range(21, 42):
            for col in ["I", "BO", "CK"]:
                self.set_cell(ws, f"{col}{row}", None)

    def normalize_date(self, value):
        if isinstance(value, datetime):
            return value.strftime("%d.%m.%Y")

        text = str(value)

        if "-" in text:
            try:
                return datetime.strptime(
                    text[:10],
                    "%Y-%m-%d"
                ).strftime("%d.%m.%Y")
            except:
                pass

        return text

    def get_day(self, value):
        try:
            if "-" in str(value):
                return int(str(value).split("-")[2])
            return int(str(value).split(".")[0])
        except:
            return None

    def find_template(self, template):
        if template and os.path.exists(template):
            return template

        if template:
            name = os.path.basename(template)
            new_path = os.path.join(TEMPLATES_DIR, name)

            if os.path.exists(new_path):
                return new_path

        return template

    def get_distribution_fuel(self, inventory, start_date, end_date):
        result = {}

        if not self.db:
            return result

        rows = self.db.get_fuel_distribution(
            start_date,
            end_date
        )

        for row in rows:
            if row.get("generator") == inventory:
                result[
                    self.normalize_date(row.get("date"))
                ] = float(row.get("liters") or 0)

        return result

    def create_report(
            self,
            generator,
            year,
            month,
            decade,
            work_data=None,
            fuel_data=None
        ):

        if not isinstance(generator, dict):
            generator = dict(generator)

        template = self.find_template(
            generator.get("template")
        )

        if not template or not os.path.exists(template):
            raise Exception(
                "Не найден шаблон Excel: " + str(template)
            )

        folder = os.path.join(
            self.reports_folder,
            str(year),
            calendar.month_name[month]
        )

        os.makedirs(folder, exist_ok=True)

        filename = (
            f"{generator.get('inventory','Генератор')}_"
            f"{self.get_decade_sheet(decade)}.xlsx"
        )

        report_file = os.path.join(
            folder,
            filename
        )

        shutil.copy(
            template,
            report_file
        )

        wb = load_workbook(report_file)

        ws = wb[
            self.get_decade_sheet(decade)
        ]

        self.clear_rows(ws)

        fuel_map = {}

        if self.db:
            fuel_map = self.get_distribution_fuel(
                generator.get("inventory"),
                f"{year}-{month:02d}-01",
                f"{year}-{month:02d}-31"
            )

        elif fuel_data:
            for k, v in fuel_data.items():
                fuel_map[self.normalize_date(k)] = float(v or 0)

        row = 21

        for item in work_data or []:

            day = self.get_day(item.get("date"))

            if day is None:
                continue

            if decade == 1 and day > 10:
                continue

            if decade == 2 and (day < 11 or day > 20):
                continue

            if decade == 3 and day < 21:
                continue

            dt = self.normalize_date(
                item.get("date")
            )

            received = fuel_map.get(
                dt,
                0
            )

            hours = float(
                item.get("hours", 0) or 0
            )

            if hours <= 0 and received <= 0:
                continue

            self.set_cell(
                ws,
                f"I{row}",
                dt
            )

            self.set_cell(
                ws,
                f"BO{row}",
                hours
            )

            self.set_cell(
                ws,
                f"CK{row}",
                received
            )

            row += 2

            if row > 41:
                break

        wb.save(report_file)

        return report_file
