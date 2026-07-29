from PySide6.QtGui import QColor, QFont

from PySide6.QtWidgets import (
    QMainWindow,
    QWidget,
    QHBoxLayout,
    QVBoxLayout,
    QLabel,
    QListWidget,
    QPushButton,
    QToolBar,
    QStatusBar,
    QGroupBox,
    QMessageBox
)

from app.ui.generator_dialog import GeneratorDialog
from app.ui.fuel_window import FuelWindow
from app.ui.report_dialog import ReportDialog
from app.services.excel_report import ExcelReport
from app.ui.work_calendar_window import WorkCalendarWindow

import os
import glob
from openpyxl import load_workbook


class MainWindow(QMainWindow):

    def __init__(self, database):

        super().__init__()

        self.db = database
        self.fuel_window = None
        self.current_generator = None

        self.setWindowTitle(
            "Generator Report Pro"
        )

        self.resize(
            1400,
            800
        )

        self.create_menu()
        self.create_toolbar()
        self.create_ui()
        self.create_statusbar()

        self.btn_add.clicked.connect(
            self.add_generator
        )

        self.btn_edit.clicked.connect(
            self.edit_generator
        )

        self.btn_delete.clicked.connect(
            self.archive_generator
        )

        self.btn_refresh.clicked.connect(
            self.load_generators
        )

        self.btn_report.clicked.connect(
            self.generate_excel_report
        )

        self.btn_calendar.clicked.connect(
            self.open_work_calendar
        )

        self.btn_up.clicked.connect(
            lambda: self.move_selected_generator(-1)
        )

        self.btn_down.clicked.connect(
            lambda: self.move_selected_generator(1)
        )

        self.generator_list.currentItemChanged.connect(
            self.show_generator
        )

        self.load_generators()



    def create_menu(self):

        menu = self.menuBar()

        menu.addMenu("Файл")
        menu.addMenu("Генераторы")
        menu.addMenu("Отчеты")


        fuel_menu = menu.addMenu(
            "Топливо"
        )

        fuel_action = fuel_menu.addAction(
            "Получение топлива"
        )

        fuel_action.triggered.connect(
            self.open_fuel_window
        )

        menu.addMenu(
            "Настройки"
        )



    def open_fuel_window(self):

        self.fuel_window = FuelWindow(
            self.db
        )

        self.fuel_window.show()



    def create_toolbar(self):

        toolbar = QToolBar()

        self.addToolBar(
            toolbar
        )


        self.btn_add = QPushButton(
            "➕ Добавить"
        )

        self.btn_edit = QPushButton(
            "✏ Редактировать"
        )

        self.btn_delete = QPushButton(
            "🗑 Архив"
        )

        self.btn_refresh = QPushButton(
            "🔄 Обновить"
        )

        self.btn_report = QPushButton(
            "📄 Отчёт Excel"
        )

        self.btn_calendar = QPushButton(
            "📅 Работа генераторов"
        )

        self.btn_up = QPushButton(
            "⬆ Вверх"
        )

        self.btn_down = QPushButton(
            "⬇ Вниз"
        )


        for btn in [
            self.btn_add,
            self.btn_edit,
            self.btn_delete,
            self.btn_refresh,
            self.btn_report,
            self.btn_calendar,
            self.btn_up,
            self.btn_down
        ]:

            toolbar.addWidget(
                btn
            )
    def create_ui(self):

        central = QWidget()

        self.setCentralWidget(
            central
        )


        main = QHBoxLayout(
            central
        )


        left = QVBoxLayout()

        left.addWidget(
            QLabel("Генераторы")
        )


        self.generator_list = QListWidget()

        left.addWidget(
            self.generator_list
        )



        box = QGroupBox(
            "Информация"
        )

        info = QVBoxLayout(
            box
        )


        self.lbl_inventory = QLabel()
        self.lbl_name = QLabel()
        self.lbl_department = QLabel()
        self.lbl_type = QLabel()
        self.lbl_model = QLabel()
        self.lbl_serial = QLabel()
        self.lbl_location = QLabel()
        self.lbl_consumption = QLabel()
        self.lbl_template = QLabel()
        self.lbl_note = QLabel()
        self.lbl_status = QLabel()
        self.lbl_balance = QLabel()


        for w in [
            self.lbl_inventory,
            self.lbl_name,
            self.lbl_department,
            self.lbl_type,
            self.lbl_model,
            self.lbl_serial,
            self.lbl_location,
            self.lbl_consumption,
            self.lbl_template,
            self.lbl_note,
            self.lbl_status,
            self.lbl_balance
        ]:
            info.addWidget(w)



        right = QVBoxLayout()

        right.addWidget(
            box
        )


        main.addLayout(
            left,
            1
        )

        main.addLayout(
            right,
            3
        )



    def create_statusbar(self):

        self.setStatusBar(
            QStatusBar()
        )



    def load_generators(self):

        self.generator_list.clear()


        for g in self.db.get_generators():

            self.generator_list.addItem(g["inventory"])

        # оформление статусов в списке
        for row, g in enumerate(self.db.get_generators()):
            item = self.generator_list.item(row)
            status = g.get("status", "Рабочий")

            if status == "Нерабочий":
                item.setForeground(QColor("red"))
                font = QFont()
                font.setBold(True)
                item.setFont(font)


        if self.generator_list.count():

            self.generator_list.setCurrentRow(
                0
            )



    def show_generator(self, item):

        if not item:
            return


        g = self.db.get_generator(
            item.text()
        )


        if not g:
            return


        self.current_generator = g


        self.lbl_inventory.setText(
            f"Инвентарный номер: {g['inventory']}"
        )

        self.lbl_name.setText(
            f"Название: {g['name'] or ''}"
        )

        self.lbl_department.setText(
            f"Подразделение: {g['department'] or ''}"
        )

        self.lbl_type.setText(
            f"Тип: {g['generator_type'] or ''}"
        )

        self.lbl_model.setText(
            f"Модель: {g['model'] or ''}"
        )

        self.lbl_serial.setText(
            f"Серийный номер: {g['serial'] or ''}"
        )

        self.lbl_location.setText(
            f"Объект: {g['location'] or ''}"
        )

        self.lbl_consumption.setText(
            f"Расход: {g['consumption'] or 0}"
        )

        self.lbl_template.setText(
            f"Шаблон: {g['template'] or ''}"
        )

        self.lbl_note.setText(
            f"Примечание: {g['note'] or ''}"
        )

        # Остаток топлива на начало:
        # сначала берём ручное значение из карточки,
        # если оно не задано — читаем из Excel
        manual_balance = g.get("initial_balance", 0) or 0

        if manual_balance > 0:
            balance = manual_balance
        else:
            balance = self.get_excel_start_balance(g)

        self.lbl_balance.setText(
            f"Остаток топлива на начало: {balance} л"
        )

        status = g.get("status", "Рабочий")
        self.lbl_status.setText(
            f"Статус: {status}"
        )
        self.lbl_status.setStyleSheet(
            "font-weight: bold; color: red;" if status == "Нерабочий"
            else "font-weight: bold;"
        )






    def get_excel_start_balance(self, generator):
        """
        Получение остатка топлива на начало из последнего Excel отчета.
        Ищет строку "Остаток на начало, л." во всех ячейках.
        """
        try:
            template = generator.get("template")

            files = []

            if template and os.path.exists(template):
                files.append(template)

            # ищем отчеты Excel рядом с проектом
            files += glob.glob("**/*.xlsx", recursive=True)

            if not files:
                return 0

            filename = max(
                files,
                key=os.path.getmtime
            )

            wb = load_workbook(filename, data_only=True)

            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and "Остаток на начало" in str(cell.value):
                            # ищем число в следующих ячейках строки
                            for c in range(cell.column + 1, ws.max_column + 1):
                                value = ws.cell(cell.row, c).value
                                if isinstance(value, (int, float)):
                                    return float(value)

        except Exception:
            pass

        return 0


    def generate_excel_report(self):

        import calendar
        from datetime import date


        if not self.current_generator:

            QMessageBox.warning(
                self,
                "Ошибка",
                "Выберите генератор"
            )

            return


        dialog = ReportDialog(self)


        if dialog.exec():


            try:

                year = dialog.year()
                month = dialog.month()
                decade = dialog.decade()


                if decade == 1:

                    start_day = 1
                    end_day = 10


                elif decade == 2:

                    start_day = 11
                    end_day = 20


                else:

                    start_day = 21

                    end_day = calendar.monthrange(
                        year,
                        month
                    )[1]


                start_date = date(
                    year,
                    month,
                    start_day
                ).isoformat()


                end_date = date(
                    year,
                    month,
                    end_day
                ).isoformat()



            # получаем реальные данные из базы

                work_data = self.db.get_daily_report(

                    self.current_generator["inventory"],

                    start_date,

                    end_date

                )



                # получаем данные получения топлива за период
                fuel_data = {}

                if hasattr(self.db, "get_fuel_income"):
                    fuel_rows = self.db.get_fuel_income()

                    for fuel in fuel_rows:
                        if not isinstance(fuel, dict):
                            fuel = dict(fuel)

                        fuel_date = fuel.get("date")
                        liters = fuel.get("liters", 0)

                        if fuel_date:
                            fuel_data[fuel_date] = float(liters or 0)


                report = ExcelReport()



                filename = report.create_report(

                    self.current_generator,

                    year,

                    month,

                    decade,

                    work_data,

                    fuel_data

                )



                QMessageBox.information(

                    self,

                    "Готово",

                    f"Отчёт создан:\n{filename}"

                )


            except Exception as e:


                QMessageBox.critical(

                    self,

                    "Ошибка",

                    str(e)

                )

    def open_work_calendar(self):

        dialog = WorkCalendarWindow(
            self.db,
            self
        )

        dialog.exec()


    def add_generator(self):

        dialog = GeneratorDialog(
            self
        )


        if dialog.exec():

            self.db.add_generator(
                dialog.get_data()
            )

            self.load_generators()



    def edit_generator(self):

        item = self.generator_list.currentItem()


        if not item:
            return


        old_inventory = item.text()

        generator = self.db.get_generator(
            old_inventory
        )


        dialog = GeneratorDialog(
            self,
            generator
        )


        if dialog.exec():

            data = dialog.get_data()


            if hasattr(
                self.db,
                "update_generator_full"
            ):

                self.db.update_generator_full(
                    old_inventory,
                    data
                )

            else:

                self.db.update_generator(
                    old_inventory,
                    data
                )


            self.load_generators()



    def move_selected_generator(self, direction):

        item = self.generator_list.currentItem()

        if not item:
            return

        self.db.move_generator(
            item.text(),
            direction
        )

        self.load_generators()


    def archive_generator(self):

        item = self.generator_list.currentItem()


        if item:

            self.db.archive_generator(
                item.text()
            )

            self.load_generators()        